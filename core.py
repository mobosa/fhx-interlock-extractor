#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
FHX Interlock Extractor - 核心模块
从DeltaV FHX文件中提取DCC功能块的联锁/允许/强制/AT模拟跟踪条件信息，生成Excel联锁信息表

关键格式说明:
  MODULE_INSTANCE条目中使用 "DCC1$参数名" 格式（$分隔）存储实际配置
  MODULE_CLASS定义中使用 "DCC1/参数名" 格式（/分隔）存储默认值
  实际联锁配置在MODULE_INSTANCE的DCC1$I_EXP/DCC1$I_DESC/DCC1$I_STATE中
"""

from dataclasses import dataclass, field
import re
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===================== 数据结构 =====================

@dataclass
class ConditionSet:
    """通用条件集合 — 联锁/允许/强制/模拟跟踪共用"""
    exps: dict[int, str] = field(default_factory=dict)
    descs: dict[int, str] = field(default_factory=dict)
    states: dict[int, str] = field(default_factory=dict)
    disable: dict[int, bool] = field(default_factory=dict)
    delay_on: dict[int, object] = field(default_factory=dict)
    delay_off: dict[int, object] = field(default_factory=dict)
    higher_mng: dict[int, bool] | None = None
    reset_reqd: dict[int, bool] | None = None
    hold_man: dict[int, bool] | None = None
    val: dict[int, object] | None = None
    nof: int = 16
    used: int = 0

    def has_real(self) -> bool:
        return any(v and not is_false_exp(v) for v in self.exps.values())

    def count_real(self) -> int:
        return sum(1 for v in self.exps.values() if v and not is_false_exp(v))

    def finalize(self):
        if self.used > 0:
            self.nof = self.used
        self.used = self.count_real()


@dataclass
class ModuleInstance:
    tag: str
    plant_area: str
    description: str
    module_class: str
    interlock: ConditionSet = field(default_factory=ConditionSet)
    permissive: ConditionSet = field(default_factory=ConditionSet)
    force: ConditionSet = field(default_factory=ConditionSet)
    tracking: ConditionSet = field(default_factory=ConditionSet)
    block_type: str = ''

    @property
    def dcc_config(self):
        """兼容旧代码访问方式"""
        return {
            'i_exps': self.interlock.exps, 'i_descs': self.interlock.descs,
            'i_states': self.interlock.states, 'i_disable': self.interlock.disable,
            'i_higher_mng': self.interlock.higher_mng or {},
            'i_reset_reqd': self.interlock.reset_reqd or {},
            'i_delay_on': self.interlock.delay_on, 'i_delay_off': self.interlock.delay_off,
            'i_used': self.interlock.used, 'i_nof': self.interlock.nof,
            'p_exps': self.permissive.exps, 'p_descs': self.permissive.descs,
            'p_disable': self.permissive.disable,
            'p_delay_on': self.permissive.delay_on, 'p_delay_off': self.permissive.delay_off,
            'p_used': self.permissive.used, 'p_nof': self.permissive.nof,
            'f_exps': self.force.exps, 'f_descs': self.force.descs,
            'f_states': self.force.states, 'f_disable': self.force.disable,
            'f_delay_on': self.force.delay_on,
            'f_used': self.force.used, 'f_nof': self.force.nof,
            't_exps': self.tracking.exps, 't_descs': self.tracking.descs,
            't_states': self.tracking.states, 't_disable': self.tracking.disable,
            't_higher_mng': self.tracking.higher_mng or {},
            't_reset_reqd': self.tracking.reset_reqd or {},
            't_hold_man': self.tracking.hold_man or {},
            't_delay_on': self.tracking.delay_on, 't_delay_off': self.tracking.delay_off,
            't_val': self.tracking.val or {},
            't_used': self.tracking.used, 't_nof': self.tracking.nof,
            'block_type': self.block_type,
        }


# ===================== 工具函数 =====================

def is_false_exp(v: str) -> bool:
    return v in ('FALSE;', 'FALSE')


def parse_bool(val) -> bool:
    return str(val).upper() in ('TRUE', 'T')


def parse_int_safe(val, default: int = 0) -> int:
    if isinstance(val, int):
        return val
    s = str(val).strip()
    if s.isdigit():
        return int(s)
    return default


# ===================== 文件读取 =====================

def read_fhx_file(filepath: str) -> str:
    with open(filepath, 'rb') as f:
        raw = f.read()
    if raw[:2] == b'\xff\xfe':
        return raw.decode('utf-16-le').replace('\x00', '')
    if raw[:2] == b'\xfe\xff':
        return raw.decode('utf-16-be').replace('\x00', '')
    # 无 BOM — 优先尝试 UTF-8，失败降级 UTF-16
    try:
        text = raw.decode('utf-8')
        text = text.replace('\x00', '')
        return text
    except UnicodeDecodeError:
        return raw.decode('utf-16', errors='ignore').replace('\x00', '')


def extract_value(lines: list[str], start: int):
    for i in range(start, min(start + 15, len(lines))):
        line = lines[i].strip()
        # CV="..." 字符串值
        cv = re.search(r'CV="([^"]*)"', line)
        if cv:
            return cv.group(1)
        # CV=N 整数值
        cv_int = re.search(r'CV=(\d+)', line)
        if cv_int:
            return int(cv_int.group(1))
        # CV=T / CV=F 布尔值
        cv_bool = re.search(r'CV=(T|F)\b', line)
        if cv_bool:
            return cv_bool.group(1)
        # EXPRESSION="..." — 可能跨多行
        exp_marker = 'EXPRESSION="'
        exp_pos = line.find(exp_marker)
        if exp_pos >= 0:
            after = line[exp_pos + len(exp_marker):]
            if '"' in after:
                last_quote = after.rfind('"')
                return after[:last_quote] if last_quote >= 0 else after
            # 多行表达式 — 续读至闭合引号
            full_val = after
            for j in range(i + 1, min(i + 10, len(lines))):
                next_line = lines[j].strip()
                full_val += '\n' + next_line
                if '"' in next_line:
                    quote_pos = next_line.rfind('"')
                    before_quote = next_line[:quote_pos]
                    full_val = full_val[:full_val.rfind('\n' + next_line)] + '\n' + before_quote
                    return full_val.strip()
            return full_val.strip()
        # STRING_VALUE="..."
        sv = re.search(r'STRING_VALUE="([^"]*)"', line)
        if sv:
            return sv.group(1)
    return None


# ===================== FHX 解析 =====================

# 属性名后缀 → (ConditionSet 字段名, 类型)
# EXP/STATE: 无下划线后缀 (I_EXP1, I_STATE1)
# DESC/DISABLE/HIGHER_MNG 等: 有下划线 (I_DESC_1, I_DISABLE_1)
_ATTR_SUFFIX_MAP = {
    'EXP':        'exps',
    'DESC_':      'descs',
    'STATE':      'states',
    'DISABLE':    'disable',
    'HIGHER_MNG': 'higher_mng',
    'RESET_REQD': 'reset_reqd',
    'HOLD_MAN':   'hold_man',
    'DELAY_ON':   'delay_on',
    'DELAY_OFF':  'delay_off',
    'VAL':        'val',
}

# 每种条件类型的默认字段配置
_DEFAULTS = {
    'I': {'higher_mng': {}, 'reset_reqd': {}},
    'P': {},
    'F': {},
    'T': {'higher_mng': {}, 'reset_reqd': {}, 'hold_man': {}, 'val': {}},
}
_NOF_DEFAULTS = {'I': 16, 'P': 8, 'F': 8, 'T': 16}


def _make_condition_sets() -> dict[str, ConditionSet]:
    sets = {}
    for prefix in ('I', 'P', 'F', 'T'):
        cs = ConditionSet(nof=_NOF_DEFAULTS[prefix])
        for attr, val in _DEFAULTS[prefix].items():
            setattr(cs, attr, dict(val))
        sets[prefix] = cs
    return sets


def _parse_attribute(attr_name: str, val, cs: ConditionSet):
    """统一属性解析 — 150 行重复代码压缩为 ~20 行"""
    for suffix, field_name in _ATTR_SUFFIX_MAP.items():
        if attr_name.startswith(suffix):
            num_m = re.search(r'(\d+)$', attr_name)
            if num_m:
                num = int(num_m.group(1))
                d = getattr(cs, field_name)
                if d is None:
                    d = {}
                    setattr(cs, field_name, d)
                if field_name in ('disable', 'higher_mng', 'reset_reqd', 'hold_man'):
                    d[num] = parse_bool(val)
                elif field_name in ('exps', 'descs', 'states'):
                    d[num] = str(val)
                else:
                    d[num] = val
            return True
    if attr_name == 'USED_CND':
        cs.used = parse_int_safe(val)
        return True
    if attr_name == 'NOF_CND':
        cs.nof = parse_int_safe(val, cs.nof)
        return True
    return False


def _parse_block(tag: str, plant_area: str, mc_name: str, block_lines: list[str]) -> ModuleInstance | None:
    desc = ''
    for bl in block_lines:
        dm = re.search(r'DESCRIPTION="([^"]*)"', bl.strip())
        if dm:
            desc = dm.group(1)
            break

    sets = _make_condition_sets()
    block_type = ''
    has_config = False

    for bi, bl in enumerate(block_lines):
        stripped = bl.strip()
        if 'ATTRIBUTE_INSTANCE' not in stripped:
            continue

        # DCC1$/AT$/SM_AT$/HM_AT$ — $分隔的实例属性
        attr_match = re.search(r'ATTRIBUTE_INSTANCE\s+NAME="(DCC1|AT\d+|SM_AT\d+|HM_AT\d+)\$(.+)"', stripped)
        if not attr_match:
            # DCC1/SM_AT1 等 — /分隔的 NOF_CND 属性
            slash_match = re.search(r'ATTRIBUTE_INSTANCE\s+NAME="(DCC1|AT\d+|SM_AT\d+|HM_AT\d+)/((?:I|P|F|T)_NOF_CND)"', stripped)
            if slash_match:
                nof_val = extract_value(block_lines, bi + 1)
                if nof_val is not None:
                    nof_type = slash_match.group(2)[0].lower()
                    cs = sets[nof_type.upper()]
                    parsed = parse_int_safe(nof_val, cs.nof)
                    if parsed > 0:
                        cs.nof = parsed
                    has_config = True
            continue

        attr_name = attr_match.group(2)
        val = extract_value(block_lines, bi + 1)
        if val is None:
            continue

        has_config = True

        # 判断块类型
        if attr_name.startswith(('I_', 'P_', 'F_')):
            block_type = 'DCC'
        elif attr_name.startswith('T_'):
            block_type = 'AT'

        # 识别前缀并分发到对应 ConditionSet
        prefix = attr_name[0] if attr_name[0] in _DEFAULTS else None
        if prefix:
            _parse_attribute(attr_name[2:], val, sets[prefix])

    if not has_config:
        return None

    # 修正 nof/used：USED_CND 是分配槽总数，实际 used 从非 FALSE 表达式计数
    for cs in sets.values():
        cs.finalize()

    # 只保留有实际条件配置的模块
    if not any(cs.has_real() for cs in sets.values()):
        return None

    return ModuleInstance(
        tag=tag, plant_area=plant_area, description=desc,
        module_class=mc_name, block_type=block_type,
        interlock=sets['I'], permissive=sets['P'],
        force=sets['F'], tracking=sets['T'],
    )


def parse_fhx(filepath: str, progress_callback=None) -> list[ModuleInstance]:
    text = read_fhx_file(filepath)
    lines = text.split('\n')
    total_lines = len(lines)
    instances = []
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        mi_match = re.match(
            r'MODULE_INSTANCE\s+TAG="([^"]*)"\s+PLANT_AREA="([^"]*)"\s+MODULE_CLASS="([^"]*)"',
            line
        )
        if mi_match:
            tag = mi_match.group(1)
            plant_area = mi_match.group(2)
            mc_name = mi_match.group(3)

            # 找块体范围
            brace_count = 0
            block_end = i
            found_open = False
            for j in range(i, len(lines)):
                for ch in lines[j]:
                    if ch == '{':
                        brace_count += 1
                        found_open = True
                    elif ch == '}':
                        brace_count -= 1
                if found_open and brace_count <= 0:
                    block_end = j
                    break

            block_lines = lines[i:block_end + 1]
            inst = _parse_block(tag, plant_area, mc_name, block_lines)
            if inst:
                instances.append(inst)

            i = block_end + 1
            continue

        if progress_callback and i % 10000 == 0:
            progress_callback(i, total_lines)
        i += 1

    return instances


# ===================== Excel 生成 =====================

def _styles():
    """返回所有 Excel 样式对象"""
    hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    hdr_blue = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hdr_red = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    hdr_green = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
    hdr_gold = PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid')
    hdr_purple = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
    nfont = Font(name='微软雅黑', size=10)
    c_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    l_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    fill_ilock = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    fill_disabled = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    fill_perm = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    fill_force = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
    fill_at = PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid')
    return {
        'hdr_font': hdr_font, 'nfont': nfont,
        'hdr_blue': hdr_blue, 'hdr_red': hdr_red, 'hdr_green': hdr_green,
        'hdr_gold': hdr_gold, 'hdr_purple': hdr_purple,
        'c_align': c_align, 'l_align': l_align, 'border': border,
        'fill_ilock': fill_ilock, 'fill_disabled': fill_disabled,
        'fill_perm': fill_perm, 'fill_force': fill_force, 'fill_at': fill_at,
    }


def _set_header(ws, row, headers, fill, s):
    for col, (text, width) in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = s['hdr_font']; c.fill = fill; c.alignment = s['c_align']; c.border = s['border']
        ws.column_dimensions[get_column_letter(col)].width = width


def _sc(ws, row, col, value, s, font=None, align=None, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or s['nfont']; c.alignment = align or s['c_align']; c.border = s['border']
    if fill: c.fill = fill
    return c


def _highlight_disabled(ws, row, ncols, s):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = s['fill_disabled']


def _write_summary_sheet(wb, instances, s):
    ws = wb.active
    ws.title = "联锁信息总表"
    ws.merge_cells('A1:V1')
    t = ws['A1']
    t.value = 'DeltaV 联锁信息总表'
    t.font = Font(name='微软雅黑', bold=True, size=16, color='1F4E79')
    t.alignment = Alignment(horizontal='center', vertical='center')
    headers = [
        ('序号', 6), ('TAG', 14), ('Plant Area', 20), ('Description', 22), ('Module Class', 18),
        ('类型', 10), ('Condition No', 8), ('Description', 22), ('Condition Expression', 48),
        ('State/Value', 14), ('Disable', 7), ('Higher Managed', 12), ('Reset Required', 11),
        ('Delay On (s)', 10), ('Delay Off (s)', 10),
        ('Track Value', 10), ('Hold in Manual', 13),
        ('Used(I)/Max', 10), ('Used(P)/Max', 10), ('Used(F)/Max', 10), ('Used(T)/Max', 10),
    ]
    _set_header(ws, 2, headers, s['hdr_blue'], s)

    row, seq = 3, 1
    for inst in sorted(instances, key=lambda x: x.tag):
        all_conds = _collect_all_conditions(inst)
        if not all_conds:
            _write_empty_row(ws, row, seq, inst, s)
            row += 1; seq += 1
        else:
            row = _write_summary_rows(ws, row, seq, inst, all_conds, s)
            seq += 1
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:V{row - 1}'


def _collect_all_conditions(inst: ModuleInstance):
    dc = inst.dcc_config
    conds = []
    # 联锁
    for num in sorted(inst.interlock.exps.keys()):
        exp = inst.interlock.exps[num]
        if exp and not is_false_exp(exp):
            conds.append(('联锁', num, inst.interlock.descs.get(num, ''), exp,
                          inst.interlock.states.get(num, ''),
                          inst.interlock.disable.get(num, False),
                          (inst.interlock.higher_mng or {}).get(num, False),
                          (inst.interlock.reset_reqd or {}).get(num, False),
                          inst.interlock.delay_on.get(num, 0),
                          inst.interlock.delay_off.get(num, 0),
                          None, None))
    # 允许
    for num in sorted(inst.permissive.exps.keys()):
        exp = inst.permissive.exps[num]
        if exp and not is_false_exp(exp):
            conds.append(('允许', num, inst.permissive.descs.get(num, ''), exp, '',
                          inst.permissive.disable.get(num, False), False, False,
                          inst.permissive.delay_on.get(num, 0),
                          inst.permissive.delay_off.get(num, 0),
                          None, None))
    # 强制
    for num in sorted(inst.force.exps.keys()):
        exp = inst.force.exps[num]
        if exp and not is_false_exp(exp):
            conds.append(('强制', num, inst.force.descs.get(num, ''), exp,
                          inst.force.states.get(num, ''),
                          inst.force.disable.get(num, False), False, False,
                          inst.force.delay_on.get(num, 0), 0,
                          None, None))
    # 模拟跟踪
    for num in sorted(inst.tracking.exps.keys()):
        exp = inst.tracking.exps[num]
        if exp and not is_false_exp(exp):
            conds.append(('模拟跟踪', num, inst.tracking.descs.get(num, ''), exp, '',
                          inst.tracking.disable.get(num, False),
                          (inst.tracking.higher_mng or {}).get(num, False),
                          (inst.tracking.reset_reqd or {}).get(num, False),
                          inst.tracking.delay_on.get(num, 0),
                          inst.tracking.delay_off.get(num, 0),
                          (inst.tracking.val or {}).get(num, 0),
                          (inst.tracking.hold_man or {}).get(num, False)))
    return conds


def _write_empty_row(ws, row, seq, inst, s):
    for col, val, align in [
        (1, seq, None), (2, inst.tag, None), (3, inst.plant_area, s['l_align']),
        (4, inst.description, s['l_align']), (5, inst.module_class, s['l_align']),
        (6, '-', None), (7, '-', None), (8, '(无配置)', s['l_align']),
        (9, '-', s['l_align']), (10, '-', None),
    ]:
        _sc(ws, row, col, val, s, align=align)
    for c in range(11, 22): _sc(ws, row, c, '', s)


def _write_summary_rows(ws, row, seq, inst, all_conds, s):
    dc = inst.dcc_config
    first = True
    seen_types = set()
    cond_seq = 0
    fill_map = {'联锁': s['fill_ilock'], '允许': s['fill_perm'], '强制': s['fill_force']}
    for ctype, cnum, cdesc, cexp, cstate, cdisable, chigher, creset, cdelay_on, cdelay_off, ctrack_val, chold_man in all_conds:
        cond_seq += 1
        _sc(ws, row, 1, seq if first else '', s)
        _sc(ws, row, 2, inst.tag if first else '', s)
        _sc(ws, row, 3, inst.plant_area if first else '', s, align=s['l_align'])
        _sc(ws, row, 4, inst.description if first else '', s, align=s['l_align'])
        _sc(ws, row, 5, inst.module_class if first else '', s, align=s['l_align'])
        _sc(ws, row, 6, ctype, s, fill=fill_map.get(ctype, s['fill_at']))
        _sc(ws, row, 7, cond_seq, s)
        _sc(ws, row, 8, cdesc, s, align=s['l_align'])
        _sc(ws, row, 9, cexp, s, align=s['l_align'])
        # State/Value
        if ctype in ('联锁', '强制'):
            _sc(ws, row, 10, cstate, s)
        else:
            _sc(ws, row, 10, '', s)
        _sc(ws, row, 11, '是' if cdisable else '否', s)
        _sc(ws, row, 12, '是' if chigher else '否', s)
        _sc(ws, row, 13, '是' if creset else '否', s)
        _sc(ws, row, 14, cdelay_on, s)
        _sc(ws, row, 15, cdelay_off, s)
        _sc(ws, row, 16, ctrack_val if ctype == '模拟跟踪' and ctrack_val is not None else '', s)
        _sc(ws, row, 17, '是' if ctype == '模拟跟踪' and chold_man else '否' if ctype == '模拟跟踪' else '', s)
        # Used/Max（每种类型只在首次出现行显示）
        for ct_check, col_idx, val in [
            ('联锁', 18, f"{dc['i_used']}/{dc['i_nof']}"),
            ('允许', 19, f"{dc['p_used']}/{dc['p_nof']}"),
            ('强制', 20, f"{dc['f_used']}/{dc['f_nof']}"),
            ('模拟跟踪', 21, f"{dc['t_used']}/{dc['t_nof']}"),
        ]:
            _sc(ws, row, col_idx, val if ctype == ct_check and ctype not in seen_types else '', s)
        seen_types.add(ctype)
        if cdisable:
            _highlight_disabled(ws, row, 21, s)
        row += 1
        first = False
    return row


def _write_detail_sheet(wb, name, title, color, headers, fill_color, write_row_fn, instances):
    ws = wb.create_sheet(name)
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    ws['A1'].value = title
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=14, color=color)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    s = _styles()
    _set_header(ws, 2, headers, fill_color, s)
    row, seq = 3, 1
    for inst in sorted(instances, key=lambda x: x.tag):
        row, seq = write_row_fn(ws, row, seq, inst, s)
    ws.freeze_panes = 'A3'
    if row > 3:
        ws.auto_filter.ref = f'A2:{get_column_letter(len(headers))}{row - 1}'


def _write_interlock_detail_row(ws, row, seq, inst, s):
    cs = inst.interlock
    for num in sorted(cs.exps.keys()):
        exp = cs.exps[num]
        if exp and not is_false_exp(exp):
            _sc(ws, row, 1, seq, s); _sc(ws, row, 2, inst.tag, s)
            _sc(ws, row, 3, inst.plant_area, s, align=s['l_align'])
            _sc(ws, row, 4, num, s); _sc(ws, row, 5, cs.descs.get(num, ''), s, align=s['l_align'])
            _sc(ws, row, 6, exp, s, align=s['l_align']); _sc(ws, row, 7, cs.states.get(num, ''), s)
            _sc(ws, row, 8, '是' if cs.disable.get(num, False) else '否', s)
            _sc(ws, row, 9, '是' if (cs.higher_mng or {}).get(num, False) else '否', s)
            _sc(ws, row, 10, '是' if (cs.reset_reqd or {}).get(num, False) else '否', s)
            _sc(ws, row, 11, cs.delay_on.get(num, 0), s)
            _sc(ws, row, 12, cs.delay_off.get(num, 0), s)
            _sc(ws, row, 13, f"{cs.used}/{cs.nof}", s)
            if cs.disable.get(num, False):
                _highlight_disabled(ws, row, 13, s)
            row += 1; seq += 1
    return row, seq


def _write_permissive_detail_row(ws, row, seq, inst, s):
    cs = inst.permissive
    for num in sorted(cs.exps.keys()):
        exp = cs.exps[num]
        if exp and not is_false_exp(exp):
            _sc(ws, row, 1, seq, s); _sc(ws, row, 2, inst.tag, s)
            _sc(ws, row, 3, inst.plant_area, s, align=s['l_align'])
            _sc(ws, row, 4, num, s); _sc(ws, row, 5, cs.descs.get(num, ''), s, align=s['l_align'])
            _sc(ws, row, 6, exp, s, align=s['l_align'])
            _sc(ws, row, 7, '是' if cs.disable.get(num, False) else '否', s)
            _sc(ws, row, 8, cs.delay_on.get(num, 0), s)
            _sc(ws, row, 9, cs.delay_off.get(num, 0), s)
            _sc(ws, row, 10, f"{cs.used}/{cs.nof}", s)
            if cs.disable.get(num, False):
                _highlight_disabled(ws, row, 10, s)
            row += 1; seq += 1
    return row, seq


def _write_force_detail_row(ws, row, seq, inst, s):
    cs = inst.force
    for num in sorted(cs.exps.keys()):
        exp = cs.exps[num]
        if exp and not is_false_exp(exp):
            _sc(ws, row, 1, seq, s); _sc(ws, row, 2, inst.tag, s)
            _sc(ws, row, 3, inst.plant_area, s, align=s['l_align'])
            _sc(ws, row, 4, num, s); _sc(ws, row, 5, cs.descs.get(num, ''), s, align=s['l_align'])
            _sc(ws, row, 6, exp, s, align=s['l_align']); _sc(ws, row, 7, cs.states.get(num, ''), s)
            _sc(ws, row, 8, '是' if cs.disable.get(num, False) else '否', s)
            _sc(ws, row, 9, cs.delay_on.get(num, 0), s)
            _sc(ws, row, 10, f"{cs.used}/{cs.nof}", s)
            if cs.disable.get(num, False):
                _highlight_disabled(ws, row, 10, s)
            row += 1; seq += 1
    return row, seq


def _write_at_detail_row(ws, row, seq, inst, s):
    cs = inst.tracking
    for num in sorted(cs.exps.keys()):
        exp = cs.exps[num]
        if exp and not is_false_exp(exp):
            _sc(ws, row, 1, seq, s); _sc(ws, row, 2, inst.tag, s)
            _sc(ws, row, 3, inst.plant_area, s, align=s['l_align'])
            _sc(ws, row, 4, num, s); _sc(ws, row, 5, cs.descs.get(num, ''), s, align=s['l_align'])
            _sc(ws, row, 6, exp, s, align=s['l_align'])
            _sc(ws, row, 7, (cs.val or {}).get(num, 0), s)
            _sc(ws, row, 8, '是' if cs.disable.get(num, False) else '否', s)
            _sc(ws, row, 9, '是' if (cs.higher_mng or {}).get(num, False) else '否', s)
            _sc(ws, row, 10, '是' if (cs.hold_man or {}).get(num, False) else '否', s)
            _sc(ws, row, 11, '是' if (cs.reset_reqd or {}).get(num, False) else '否', s)
            _sc(ws, row, 12, cs.delay_on.get(num, 0), s)
            _sc(ws, row, 13, cs.delay_off.get(num, 0), s)
            _sc(ws, row, 14, f"{cs.used}/{cs.nof}", s)
            if cs.disable.get(num, False):
                _highlight_disabled(ws, row, 14, s)
            row += 1; seq += 1
    return row, seq


def _write_module_summary(wb, instances):
    ws = wb.create_sheet("模块实例汇总")
    ws.merge_cells('A1:I1')
    ws['A1'].value = '含联锁配置的模块实例汇总'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=14, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    s = _styles()
    h5 = [('序号', 6), ('TAG', 14), ('Plant Area', 20), ('Description', 22), ('Module Class', 18),
          ('联锁数', 10), ('允许数', 10), ('强制数', 10), ('模拟跟踪数', 10)]
    _set_header(ws, 2, h5, s['hdr_blue'], s)
    row, seq = 3, 1
    for inst in sorted(instances, key=lambda x: x.tag):
        for col, val in [
            (1, seq), (2, inst.tag), (3, inst.plant_area), (4, inst.description),
            (5, inst.module_class),
            (6, inst.interlock.count_real()), (7, inst.permissive.count_real()),
            (8, inst.force.count_real()), (9, inst.tracking.count_real()),
        ]:
            align = s['l_align'] if col in (3, 4, 5) else s['c_align']
            _sc(ws, row, col, val, s, align=align)
        row += 1; seq += 1
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:I{row - 1}'


def generate_excel(instances: list[ModuleInstance], output_path: str) -> str:
    wb = Workbook()
    s = _styles()

    _write_summary_sheet(wb, instances, s)

    _write_detail_sheet(wb, "联锁条件明细", '联锁条件明细表', 'C00000',
        [('序号', 6), ('TAG', 14), ('Plant Area', 18), ('Condition No', 8), ('Description', 22),
         ('Condition Expression', 48), ('Interlock State', 12), ('Disable', 7),
         ('Higher Managed', 12), ('Reset Required', 11), ('Delay On (s)', 10),
         ('Delay Off (s)', 10), ('Used/Max', 10)],
        s['hdr_red'], _write_interlock_detail_row, instances)

    _write_detail_sheet(wb, "允许条件明细", '允许条件明细表', '548235',
        [('序号', 6), ('TAG', 14), ('Plant Area', 18), ('Condition No', 8), ('Description', 22),
         ('Condition Expression', 48), ('Disable', 7), ('Delay On (s)', 10),
         ('Delay Off (s)', 10), ('Used/Max', 10)],
        s['hdr_green'], _write_permissive_detail_row, instances)

    _write_detail_sheet(wb, "强制条件明细", '强制设定点条件明细表', 'BF8F00',
        [('序号', 6), ('TAG', 14), ('Plant Area', 18), ('Condition No', 8), ('Description', 22),
         ('Condition Expression', 48), ('Force SP Value', 12), ('Disable', 7),
         ('Delay On (s)', 10), ('Used/Max', 10)],
        s['hdr_gold'], _write_force_detail_row, instances)

    _write_detail_sheet(wb, "模拟跟踪条件明细", 'Analog Tracking (AT) 条件明细表', '7030A0',
        [('序号', 6), ('TAG', 14), ('Plant Area', 18), ('Condition No', 8), ('Description', 22),
         ('Condition Expression', 48), ('Track Value', 10), ('Disable', 7),
         ('Higher Managed', 12), ('Hold in Manual', 13), ('Reset Required', 11),
         ('Delay On (s)', 10), ('Delay Off (s)', 10), ('Used/Max', 10)],
        s['hdr_purple'], _write_at_detail_row, instances)

    _write_module_summary(wb, instances)

    saved_path = _save_workbook(wb, output_path)
    return saved_path


def _save_workbook(wb, output_path: str) -> str:
    """保存工作簿，文件被占用时给出明确提示"""
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        raise PermissionError(
            f"无法保存文件，目标文件被占用，请关闭已打开的 Excel 后重试:\n{output_path}"
        )
