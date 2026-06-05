#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
FHX Interlock Extractor GUI
DeltaV FHX文件联锁信息提取工具 - 图形界面版
"""

import re
import os
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===================== 解析核心逻辑 =====================

def read_fhx_file(filepath):
    with open(filepath, 'rb') as f:
        raw = f.read()
    if raw[:2] == b'\xff\xfe':
        text = raw.decode('utf-16-le')
    elif raw[:2] == b'\xfe\xff':
        text = raw.decode('utf-16-be')
    else:
        text = raw.decode('utf-16', errors='ignore')
    return text.replace('\x00', '')


def extract_value(lines, start):
    for i in range(start, min(start + 15, len(lines))):
        line = lines[i].strip()
        cv = re.search(r'CV="([^"]*)"', line)
        if cv:
            return cv.group(1)
        cv_int = re.search(r'CV=(\d+)', line)
        if cv_int:
            return int(cv_int.group(1))
        # 匹配 CV=T 或 CV=F 等布尔值
        cv_bool = re.search(r'CV=(T|F)\b', line)
        if cv_bool:
            return cv_bool.group(1)
        exp_marker = 'EXPRESSION="'
        exp_pos = line.find(exp_marker)
        if exp_pos >= 0:
            after = line[exp_pos + len(exp_marker):]
            if '"' in after:
                last_quote = after.rfind('"')
                if last_quote >= 0:
                    return after[:last_quote]
                return after
            else:
                full_val = after
                for j in range(i + 1, min(i + 10, len(lines))):
                    next_line = lines[j].strip()
                    full_val += '\n' + next_line
                    if '"' in next_line:
                        quote_pos = next_line.rfind('"')
                        before_quote = next_line[:quote_pos]
                        full_val = full_val[:full_val.rfind('\n' + next_line)] + '\n' + before_quote
                        return full_val.strip()
        sv = re.search(r'STRING_VALUE="([^"]*)"', line)
        if sv:
            return sv.group(1)
    return None


def parse_fhx(filepath, progress_callback=None):
    text = read_fhx_file(filepath)
    lines = text.split('\n')
    total_lines = len(lines)
    instances = []
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        mi_match = re.match(
            r'MODULE_INSTANCE\s+TAG="([^"]*)"\s+PLANT_AREA="([^"]*)"\s+MODULE_CLASS="([^"]*)"',
            line
        )
        if mi_match:
            tag = mi_match.group(1)
            plant_area = mi_match.group(2)
            mc_name = mi_match.group(3)
            brace_count = 0
            block_start = i
            block_end = i
            found_open = False
            for j in range(i, min(i + 5000, len(lines))):
                for ch in lines[j]:
                    if ch == '{':
                        brace_count += 1
                        found_open = True
                    elif ch == '}':
                        brace_count -= 1
                if found_open and brace_count <= 0:
                    block_end = j
                    break
            block_lines = lines[block_start:block_end + 1]
            desc = ''
            for bl in block_lines:
                dm = re.search(r'DESCRIPTION="([^"]*)"', bl.strip())
                if dm:
                    desc = dm.group(1)
                    break
            dcc_config = {
                'i_exps': {}, 'i_descs': {}, 'i_states': {},
                'i_disable': {}, 'i_higher_mng': {}, 'i_reset_reqd': {},
                'i_delay_on': {}, 'i_delay_off': {}, 'i_used': 0,
                'p_exps': {}, 'p_descs': {}, 'p_disable': {},
                'p_delay_on': {}, 'p_delay_off': {}, 'p_used': 0,
                'f_exps': {}, 'f_descs': {}, 'f_states': {},
                'f_disable': {}, 'f_delay_on': {}, 'f_used': 0,
                't_exps': {}, 't_descs': {}, 't_states': {},
                't_disable': {}, 't_higher_mng': {}, 't_reset_reqd': {}, 't_hold_man': {},
                't_delay_on': {}, 't_delay_off': {}, 't_val': {}, 't_used': 0,
                'block_type': '',
            }
            has_dcc_config = False
            for bi, bl in enumerate(block_lines):
                stripped = bl.strip()
                if 'ATTRIBUTE_INSTANCE' not in stripped:
                    continue
                attr_match = re.search(r'ATTRIBUTE_INSTANCE\s+NAME="(DCC1|AT\d+|SM_AT\d+|HM_AT\d+)\$(.+)"', stripped)
                if not attr_match:
                    continue
                block_name = attr_match.group(1)
                attr_name = attr_match.group(2)
                val = extract_value(block_lines, bi + 1)
                if val is None:
                    continue
                has_dcc_config = True
                if attr_name.startswith(('I_', 'P_', 'F_')):
                    dcc_config['block_type'] = 'DCC'
                elif attr_name.startswith('T_'):
                    dcc_config['block_type'] = 'AT'
                # DCC Interlock
                m = re.match(r'I_EXP(\d+)$', attr_name)
                if m:
                    dcc_config['i_exps'][int(m.group(1))] = str(val); continue
                m = re.match(r'I_DESC_(\d+)$', attr_name)
                if m:
                    dcc_config['i_descs'][int(m.group(1))] = str(val); continue
                m = re.match(r'I_STATE(\d+)$', attr_name)
                if m:
                    dcc_config['i_states'][int(m.group(1))] = str(val); continue
                if attr_name == 'I_USED_CND':
                    dcc_config['i_used'] = int(val) if isinstance(val, int) else 0; continue
                m = re.match(r'I_DISABLE(\d+)$', attr_name)
                if m:
                    dcc_config['i_disable'][int(m.group(1))] = str(val).upper() in ('TRUE', 'T'); continue
                m = re.match(r'I_HIGHER_MNG(\d+)$', attr_name)
                if m:
                    dcc_config['i_higher_mng'][int(m.group(1))] = str(val).upper() in ('TRUE', 'T'); continue
                m = re.match(r'I_RESET_REQD(\d+)$', attr_name)
                if m:
                    dcc_config['i_reset_reqd'][int(m.group(1))] = str(val).upper() in ('TRUE', 'T'); continue
                m = re.match(r'I_DELAY_ON(\d+)$', attr_name)
                if m:
                    dcc_config['i_delay_on'][int(m.group(1))] = val; continue
                m = re.match(r'I_DELAY_OFF(\d+)$', attr_name)
                if m:
                    dcc_config['i_delay_off'][int(m.group(1))] = val; continue
                # DCC Permissive
                m = re.match(r'P_EXP(\d+)$', attr_name)
                if m:
                    dcc_config['p_exps'][int(m.group(1))] = str(val); continue
                m = re.match(r'P_DESC(\d+)$', attr_name)
                if m:
                    dcc_config['p_descs'][int(m.group(1))] = str(val); continue
                if attr_name == 'P_USED_CND':
                    dcc_config['p_used'] = int(val) if isinstance(val, int) else 0; continue
                m = re.match(r'P_DISABLE(\d+)$', attr_name)
                if m:
                    dcc_config['p_disable'][int(m.group(1))] = str(val).upper() in ('TRUE', 'T'); continue
                m = re.match(r'P_DELAY_ON(\d+)$', attr_name)
                if m:
                    dcc_config['p_delay_on'][int(m.group(1))] = val; continue
                m = re.match(r'P_DELAY_OFF(\d+)$', attr_name)
                if m:
                    dcc_config['p_delay_off'][int(m.group(1))] = val; continue
                # DCC Force
                m = re.match(r'F_EXP(\d+)$', attr_name)
                if m:
                    dcc_config['f_exps'][int(m.group(1))] = str(val); continue
                m = re.match(r'F_DESC(\d+)$', attr_name)
                if m:
                    dcc_config['f_descs'][int(m.group(1))] = str(val); continue
                m = re.match(r'F_STATE(\d+)$', attr_name)
                if m:
                    dcc_config['f_states'][int(m.group(1))] = str(val); continue
                if attr_name == 'F_USED_CND':
                    dcc_config['f_used'] = int(val) if isinstance(val, int) else 0; continue
                m = re.match(r'F_DISABLE(\d+)$', attr_name)
                if m:
                    dcc_config['f_disable'][int(m.group(1))] = str(val).upper() in ('TRUE', 'T'); continue
                m = re.match(r'F_DELAY_ON(\d+)$', attr_name)
                if m:
                    dcc_config['f_delay_on'][int(m.group(1))] = val; continue
                # AT Tracking
                m = re.match(r'T_EXP(\d+)$', attr_name)
                if m:
                    dcc_config['t_exps'][int(m.group(1))] = str(val); continue
                m = re.match(r'T_DESC(\d+)$', attr_name)
                if m:
                    dcc_config['t_descs'][int(m.group(1))] = str(val); continue
                m = re.match(r'T_STATE(\d+)$', attr_name)
                if m:
                    dcc_config['t_states'][int(m.group(1))] = str(val); continue
                if attr_name == 'T_USED_CND':
                    dcc_config['t_used'] = int(val) if isinstance(val, int) else 0; continue
                m = re.match(r'T_DISABLE(\d+)$', attr_name)
                if m:
                    dcc_config['t_disable'][int(m.group(1))] = str(val).upper() in ('TRUE', 'T'); continue
                m = re.match(r'T_HIGHER_MNG(\d+)$', attr_name)
                if m:
                    dcc_config['t_higher_mng'][int(m.group(1))] = str(val).upper() in ('TRUE', 'T')
                    continue
                m = re.match(r'T_RESET_REQD(\d+)$', attr_name)
                if m:
                    dcc_config['t_reset_reqd'][int(m.group(1))] = str(val).upper() in ('TRUE', 'T'); continue
                m = re.match(r'T_HOLD_MAN(\d+)$', attr_name)
                if m:
                    dcc_config['t_hold_man'][int(m.group(1))] = str(val).upper() in ('TRUE', 'T'); continue
                m = re.match(r'T_DELAY_ON(\d+)$', attr_name)
                if m:
                    dcc_config['t_delay_on'][int(m.group(1))] = val; continue
                m = re.match(r'T_DELAY_OFF(\d+)$', attr_name)
                if m:
                    dcc_config['t_delay_off'][int(m.group(1))] = val; continue
                m = re.match(r'T_VAL(\d+)$', attr_name)
                if m:
                    dcc_config['t_val'][int(m.group(1))] = val; continue
            if has_dcc_config:
                dc = dcc_config
                has_real = any(
                    v and v not in ('FALSE;', 'FALSE')
                    for exp_dict in [dc['i_exps'], dc['p_exps'], dc['f_exps'], dc['t_exps']]
                    for v in exp_dict.values()
                )
                if has_real:
                    instances.append({
                        'tag': tag, 'plant_area': plant_area,
                        'description': desc, 'module_class': mc_name,
                        'dcc_config': dcc_config,
                    })
            i = block_end + 1
            continue
        if progress_callback and i % 10000 == 0:
            progress_callback(i, total_lines)
        i += 1
    return instances


def generate_excel(instances, output_path):
    wb = Workbook()
    hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    hdr_fill_blue = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hdr_fill_red = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    hdr_fill_green = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
    hdr_fill_gold = PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid')
    hdr_fill_purple = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
    nfont = Font(name='微软雅黑', size=10)
    c_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    l_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    fill_ilock = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    fill_perm = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    fill_force = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
    fill_at = PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid')

    def set_header(ws, row, headers, fill):
        for col, (text, width) in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=text)
            c.font = hdr_font; c.fill = fill; c.alignment = c_align; c.border = border
            ws.column_dimensions[get_column_letter(col)].width = width

    def sc(ws, row, col, value, font=nfont, align=c_align, fill=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font; c.alignment = align; c.border = border
        if fill: c.fill = fill
        return c

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "联锁信息总表"
    ws1.merge_cells('A1:S1')
    t = ws1['A1']; t.value = 'DeltaV 联锁信息总表'
    t.font = Font(name='微软雅黑', bold=True, size=16, color='1F4E79')
    t.alignment = Alignment(horizontal='center', vertical='center')
    h1 = [('序号', 6), ('TAG', 14), ('Plant Area', 20), ('Description', 22), ('Module Class', 18),
          ('类型', 10), ('Condition No', 8), ('Description', 22), ('Condition Expression', 48),
          ('Interlock State', 12), ('Force SP Value', 13), ('Disable', 7), ('Higher Managed', 12),
          ('Reset Required', 11), ('Delay On (s)', 10), ('Delay Off (s)', 10),
          ('Used(I)', 9), ('Used(P)', 9), ('Used(F)', 9)]
    set_header(ws1, 2, h1, hdr_fill_blue)
    row = 3; seq = 1
    for inst in sorted(instances, key=lambda x: x['tag']):
        dc = inst['dcc_config']
        all_conds = []
        for num in sorted(dc['i_exps'].keys()):
            exp = dc['i_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE') and not dc['i_disable'].get(num, False):
                all_conds.append(('联锁', num, dc['i_descs'].get(num, ''), exp, dc['i_states'].get(num, ''),
                                  dc['i_disable'].get(num, False), dc['i_higher_mng'].get(num, False),
                                  dc['i_reset_reqd'].get(num, False), dc['i_delay_on'].get(num, 0), dc['i_delay_off'].get(num, 0)))
        for num in sorted(dc['p_exps'].keys()):
            exp = dc['p_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE'):
                all_conds.append(('允许', num, dc['p_descs'].get(num, ''), exp, '', False, False, False, 0, 0))
        for num in sorted(dc['f_exps'].keys()):
            exp = dc['f_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE'):
                all_conds.append(('强制', num, dc['f_descs'].get(num, ''), exp, dc['f_states'].get(num, ''), False, False, False, 0, 0))
        for num in sorted(dc['t_exps'].keys()):
            exp = dc['t_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE') and not dc['t_disable'].get(num, False):
                all_conds.append(('模拟跟踪', num, dc['t_descs'].get(num, ''), exp, '',
                                  dc['t_disable'].get(num, False), False, dc['t_reset_reqd'].get(num, False),
                                  dc['t_delay_on'].get(num, 0), dc['t_delay_off'].get(num, 0)))
        if not all_conds:
            sc(ws1, row, 1, seq); sc(ws1, row, 2, inst['tag']); sc(ws1, row, 3, inst['plant_area'], align=l_align)
            sc(ws1, row, 4, inst['description'], align=l_align); sc(ws1, row, 5, inst['module_class'], align=l_align)
            sc(ws1, row, 6, '-'); sc(ws1, row, 7, '-'); sc(ws1, row, 8, '(无配置)', align=l_align)
            sc(ws1, row, 9, '-'); sc(ws1, row, 10, '-'); sc(ws1, row, 11, '-')
            for c in range(12, 20): sc(ws1, row, c, '')
            row += 1; seq += 1
        else:
            first = True
            cond_seq = 0
            for ctype, cnum, cdesc, cexp, cstate, cdisable, chigher, creset, cdelay_on, cdelay_off in all_conds:
                cond_seq += 1
                sc(ws1, row, 1, seq if first else ''); sc(ws1, row, 2, inst['tag'] if first else '')
                sc(ws1, row, 3, inst['plant_area'] if first else '', align=l_align)
                sc(ws1, row, 4, inst['description'] if first else '', align=l_align)
                sc(ws1, row, 5, inst['module_class'] if first else '', align=l_align)
                fill_map = {'联锁': fill_ilock, '允许': fill_perm, '强制': fill_force}
                sc(ws1, row, 6, ctype, fill=fill_map.get(ctype, fill_at))
                sc(ws1, row, 7, cond_seq); sc(ws1, row, 8, cdesc, align=l_align); sc(ws1, row, 9, cexp, align=l_align)
                sc(ws1, row, 10, cstate if ctype == '联锁' else '')
                sc(ws1, row, 11, cstate if ctype == '强制' else '')
                sc(ws1, row, 12, '是' if cdisable else '否')
                sc(ws1, row, 13, '是' if chigher else '否')
                sc(ws1, row, 14, '是' if creset else '否')
                sc(ws1, row, 15, cdelay_on); sc(ws1, row, 16, cdelay_off)
                sc(ws1, row, 17, dc['i_used'] if first else '')
                sc(ws1, row, 18, dc['p_used'] if first else '')
                sc(ws1, row, 19, dc['f_used'] if first else '')
                row += 1; first = False
            seq += 1
    ws1.freeze_panes = 'A3'; ws1.auto_filter.ref = f'A2:S{row - 1}'

    # Sheet 2: Interlock Detail
    ws2 = wb.create_sheet("联锁条件明细")
    ws2.merge_cells('A1:M1')
    ws2['A1'].value = '联锁条件明细表'
    ws2['A1'].font = Font(name='微软雅黑', bold=True, size=14, color='C00000')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    h2 = [('序号', 6), ('TAG', 14), ('Plant Area', 18), ('Condition No', 8), ('Description', 22),
          ('Condition Expression', 48), ('Interlock State', 12), ('Disable', 7),
          ('Higher Managed', 12), ('Reset Required', 11), ('Delay On (s)', 10),
          ('Delay Off (s)', 10), ('Used/Max', 10)]
    set_header(ws2, 2, h2, hdr_fill_red)
    row = 3; seq = 1
    for inst in sorted(instances, key=lambda x: x['tag']):
        dc = inst['dcc_config']
        cond_no = 0
        for num in sorted(dc['i_exps'].keys()):
            exp = dc['i_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE') and not dc['i_disable'].get(num, False):
                cond_no += 1
                sc(ws2, row, 1, seq); sc(ws2, row, 2, inst['tag']); sc(ws2, row, 3, inst['plant_area'], align=l_align)
                sc(ws2, row, 4, cond_no); sc(ws2, row, 5, dc['i_descs'].get(num, ''), align=l_align)
                sc(ws2, row, 6, exp, align=l_align); sc(ws2, row, 7, dc['i_states'].get(num, ''))
                sc(ws2, row, 8, '是' if dc['i_disable'].get(num, False) else '否')
                sc(ws2, row, 9, '是' if dc['i_higher_mng'].get(num, False) else '否')
                sc(ws2, row, 10, '是' if dc['i_reset_reqd'].get(num, False) else '否')
                sc(ws2, row, 11, dc['i_delay_on'].get(num, 0)); sc(ws2, row, 12, dc['i_delay_off'].get(num, 0))
                sc(ws2, row, 13, f"{dc['i_used']}/16"); row += 1; seq += 1
    ws2.freeze_panes = 'A3'
    if row > 3: ws2.auto_filter.ref = f'A2:M{row - 1}'

    # Sheet 3: Permissive Detail
    ws3 = wb.create_sheet("允许条件明细")
    ws3.merge_cells('A1:J1')
    ws3['A1'].value = '允许条件明细表'
    ws3['A1'].font = Font(name='微软雅黑', bold=True, size=14, color='548235')
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    h3 = [('序号', 6), ('TAG', 14), ('Plant Area', 18), ('Condition No', 8), ('Description', 22),
          ('Condition Expression', 48), ('Disable', 7), ('Delay On (s)', 10),
          ('Delay Off (s)', 10), ('Used/Max', 10)]
    set_header(ws3, 2, h3, hdr_fill_green)
    row = 3; seq = 1
    for inst in sorted(instances, key=lambda x: x['tag']):
        dc = inst['dcc_config']
        cond_no = 0
        for num in sorted(dc['p_exps'].keys()):
            exp = dc['p_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE'):
                cond_no += 1
                sc(ws3, row, 1, seq); sc(ws3, row, 2, inst['tag']); sc(ws3, row, 3, inst['plant_area'], align=l_align)
                sc(ws3, row, 4, cond_no); sc(ws3, row, 5, dc['p_descs'].get(num, ''), align=l_align)
                sc(ws3, row, 6, exp, align=l_align)
                sc(ws3, row, 7, '是' if dc['p_disable'].get(num, False) else '否')
                sc(ws3, row, 8, dc['p_delay_on'].get(num, 0)); sc(ws3, row, 9, dc['p_delay_off'].get(num, 0))
                sc(ws3, row, 10, f"{dc['p_used']}/8"); row += 1; seq += 1
    ws3.freeze_panes = 'A3'
    if row > 3: ws3.auto_filter.ref = f'A2:J{row - 1}'

    # Sheet 4: Force Detail
    ws4 = wb.create_sheet("强制条件明细")
    ws4.merge_cells('A1:J1')
    ws4['A1'].value = '强制设定点条件明细表'
    ws4['A1'].font = Font(name='微软雅黑', bold=True, size=14, color='BF8F00')
    ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
    h4 = [('序号', 6), ('TAG', 14), ('Plant Area', 18), ('Condition No', 8), ('Description', 22),
          ('Condition Expression', 48), ('Force SP Value', 12), ('Disable', 7),
          ('Delay On (s)', 10), ('Used/Max', 10)]
    set_header(ws4, 2, h4, hdr_fill_gold)
    row = 3; seq = 1
    for inst in sorted(instances, key=lambda x: x['tag']):
        dc = inst['dcc_config']
        cond_no = 0
        for num in sorted(dc['f_exps'].keys()):
            exp = dc['f_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE'):
                cond_no += 1
                sc(ws4, row, 1, seq); sc(ws4, row, 2, inst['tag']); sc(ws4, row, 3, inst['plant_area'], align=l_align)
                sc(ws4, row, 4, cond_no); sc(ws4, row, 5, dc['f_descs'].get(num, ''), align=l_align)
                sc(ws4, row, 6, exp, align=l_align); sc(ws4, row, 7, dc['f_states'].get(num, ''))
                sc(ws4, row, 8, '是' if dc['f_disable'].get(num, False) else '否')
                sc(ws4, row, 9, dc['f_delay_on'].get(num, 0))
                sc(ws4, row, 10, f"{dc['f_used']}/8"); row += 1; seq += 1
    ws4.freeze_panes = 'A3'
    if row > 3: ws4.auto_filter.ref = f'A2:J{row - 1}'

    # Sheet 5: AT Detail
    ws_at = wb.create_sheet("模拟跟踪条件明细")
    ws_at.merge_cells('A1:M1')
    ws_at['A1'].value = 'Analog Tracking (AT) 条件明细表'
    ws_at['A1'].font = Font(name='微软雅黑', bold=True, size=14, color='7030A0')
    ws_at['A1'].alignment = Alignment(horizontal='center', vertical='center')
    h_at = [('序号', 6), ('TAG', 14), ('Plant Area', 18), ('Condition No', 8), ('Description', 22),
            ('Condition Expression', 48), ('Track Value', 10), ('Disable', 7),
            ('Higher Managed', 12), ('Hold in Manual', 13), ('Reset Required', 11),
            ('Delay On (s)', 10), ('Delay Off (s)', 10), ('Used/Max', 10)]
    set_header(ws_at, 2, h_at, hdr_fill_purple)
    row = 3; seq = 1
    for inst in sorted(instances, key=lambda x: x['tag']):
        dc = inst['dcc_config']
        cond_no = 0
        for num in sorted(dc['t_exps'].keys()):
            exp = dc['t_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE') and not dc['t_disable'].get(num, False):
                cond_no += 1
                sc(ws_at, row, 1, seq); sc(ws_at, row, 2, inst['tag']); sc(ws_at, row, 3, inst['plant_area'], align=l_align)
                sc(ws_at, row, 4, cond_no); sc(ws_at, row, 5, dc['t_descs'].get(num, ''), align=l_align)
                sc(ws_at, row, 6, exp, align=l_align); sc(ws_at, row, 7, dc['t_val'].get(num, 0))
                sc(ws_at, row, 8, '是' if dc['t_disable'].get(num, False) else '否')
                sc(ws_at, row, 9, '是' if dc['t_higher_mng'].get(num, False) else '否')
                sc(ws_at, row, 10, '是' if dc['t_hold_man'].get(num, False) else '否')
                sc(ws_at, row, 11, '是' if dc['t_reset_reqd'].get(num, False) else '否')
                sc(ws_at, row, 12, dc['t_delay_on'].get(num, 0)); sc(ws_at, row, 13, dc['t_delay_off'].get(num, 0))
                sc(ws_at, row, 14, f"{dc['t_used']}/16"); row += 1; seq += 1
    ws_at.freeze_panes = 'A3'
    if row > 3: ws_at.auto_filter.ref = f'A2:M{row - 1}'

    # Sheet 6: Module Summary
    ws5 = wb.create_sheet("模块实例汇总")
    ws5.merge_cells('A1:G1')
    ws5['A1'].value = '含联锁配置的模块实例汇总'
    ws5['A1'].font = Font(name='微软雅黑', bold=True, size=14, color='1F4E79')
    ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
    h5 = [('序号', 6), ('TAG', 14), ('Plant Area', 20), ('Description', 22), ('Module Class', 18), ('联锁数', 10), ('允许数', 10)]
    set_header(ws5, 2, h5, hdr_fill_blue)
    row = 3; seq = 1
    for inst in sorted(instances, key=lambda x: x['tag']):
        dc = inst['dcc_config']
        ilock_count = sum(1 for n, v in dc['i_exps'].items() if v and v not in ('FALSE;', 'FALSE') and not dc['i_disable'].get(n, False))
        perm_count = sum(1 for v in dc['p_exps'].values() if v and v not in ('FALSE;', 'FALSE'))
        sc(ws5, row, 1, seq); sc(ws5, row, 2, inst['tag']); sc(ws5, row, 3, inst['plant_area'], align=l_align)
        sc(ws5, row, 4, inst['description'], align=l_align); sc(ws5, row, 5, inst['module_class'], align=l_align)
        sc(ws5, row, 6, ilock_count); sc(ws5, row, 7, perm_count)
        row += 1; seq += 1
    ws5.freeze_panes = 'A3'; ws5.auto_filter.ref = f'A2:G{row - 1}'

    wb.save(output_path)
    return output_path


# ===================== GUI =====================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('DeltaV FHX 联锁信息提取工具')
        self.geometry('600x480')
        self.resizable(False, False)
        self.configure(bg='#f0f0f0')
        self.fhx_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self._build_ui()

    def _build_ui(self):
        # Title
        title_frame = tk.Frame(self, bg='#2c3e50', height=50)
        title_frame.pack(fill='x')
        title_frame.pack_propagate(False)
        tk.Label(title_frame, text='DeltaV FHX 联锁信息提取工具',
                 font=('微软雅黑', 16, 'bold'), fg='white', bg='#2c3e50').pack(expand=True)

        # Main frame
        main = tk.Frame(self, bg='#f0f0f0', padx=20, pady=15)
        main.pack(fill='both', expand=True)

        # FHX file selection
        tk.Label(main, text='选择 FHX 文件:', font=('微软雅黑', 11, 'bold'), bg='#f0f0f0').grid(row=0, column=0, sticky='w', pady=(0, 5))
        fhx_frame = tk.Frame(main, bg='#f0f0f0')
        fhx_frame.grid(row=1, column=0, columnspan=2, sticky='ew', pady=(0, 10))
        tk.Entry(fhx_frame, textvariable=self.fhx_path, font=('微软雅黑', 10), width=58).pack(side='left', padx=(0, 5))
        tk.Button(fhx_frame, text='浏览...', command=self._browse_fhx, font=('微软雅黑', 9)).pack(side='left')

        # Output info
        self.output_var = tk.StringVar(value='输出位置: 与FHX文件同目录')
        tk.Label(main, textvariable=self.output_var, font=('微软雅黑', 9), bg='#f0f0f0', fg='#888').grid(row=2, column=0, columnspan=2, sticky='w')

        # Progress info
        self.progress_var = tk.StringVar(value='')
        self.progress_label = tk.Label(main, textvariable=self.progress_var, font=('微软雅黑', 10),
                                       bg='#27ae60', fg='white', anchor='w', padx=10, pady=4)
        self.progress_label.grid(row=3, column=0, columnspan=2, sticky='ew', pady=(10, 5))
        self.progress_label.grid_remove()  # hidden initially

        # Status label
        self.status_var = tk.StringVar(value='就绪 - 请选择FHX文件')
        tk.Label(main, textvariable=self.status_var, font=('微软雅黑', 9), bg='#f0f0f0',
                 fg='#555').grid(row=4, column=0, columnspan=2, sticky='w')

        # Run button
        self.run_btn = tk.Button(main, text='▶ 开始提取', command=self._run, font=('微软雅黑', 12, 'bold'),
                                 bg='#27ae60', fg='white', relief='flat', padx=30, pady=8,
                                 activebackground='#219a52', cursor='hand2')
        self.run_btn.grid(row=5, column=0, columnspan=2, pady=(10, 5))

        # Result text
        self.result_text = tk.Text(main, height=6, font=('Consolas', 9), bg='#ecf0f1', relief='flat',
                                   state='disabled', wrap='word')
        self.result_text.grid(row=6, column=0, columnspan=2, sticky='ew', pady=(5, 0))

        # Author info
        tk.Label(main, text='Author: Jared.Ji@emerson.com', font=('微软雅黑', 8), bg='#f0f0f0',
                 fg='#aaa').grid(row=7, column=0, columnspan=2, pady=(5, 0))

        main.columnconfigure(0, weight=1)

    def _browse_fhx(self):
        path = filedialog.askopenfilename(title='选择 FHX 文件', filetypes=[('FHX files', '*.fhx'), ('All files', '*.*')])
        if path:
            self.fhx_path.set(path)
            base = os.path.dirname(path)
            self.output_path.set(os.path.join(base, '联锁信息表.xlsx'))
            self.output_var.set(f'输出: {base}\\联锁信息表.xlsx')

    def _start_progress(self):
        self.progress_label.grid()
        self.progress_var.set('正在解析FHX文件...')

    def _stop_progress(self):
        self.progress_label.grid_remove()
        self.progress_var.set('')

    def _log(self, msg):
        self.result_text.configure(state='normal')
        self.result_text.insert('end', msg + '\n')
        self.result_text.see('end')
        self.result_text.configure(state='disabled')

    def _run(self):
        fhx = self.fhx_path.get().strip()
        if not fhx:
            messagebox.showwarning('提示', '请选择FHX文件')
            return
        if not os.path.exists(fhx):
            messagebox.showerror('错误', f'文件不存在:\n{fhx}')
            return
        out = os.path.join(os.path.dirname(fhx), '联锁信息表.xlsx')
        self.output_path.set(out)

        self.run_btn.configure(state='disabled', text='处理中...')
        self._start_progress()
        self.status_var.set('正在解析FHX文件...')
        self.result_text.configure(state='normal')
        self.result_text.delete('1.0', 'end')
        self.result_text.configure(state='disabled')

        def worker():
            try:
                self._log(f'[1/3] 读取: {os.path.basename(fhx)} ({os.path.getsize(fhx)/1024:.0f} KB)')
                self.progress_var.set('正在解析FHX文件，请稍候...')
                instances = parse_fhx(fhx)
                self.progress_var.set(f'解析完成: {len(instances)}个模块，正在统计...')
                total_ilock = sum(sum(1 for n, v in inst['dcc_config']['i_exps'].items()
                                      if v and v not in ('FALSE;', 'FALSE') and not inst['dcc_config']['i_disable'].get(n, False))
                                  for inst in instances)
                total_perm = sum(sum(1 for v in inst['dcc_config']['p_exps'].values() if v and v not in ('FALSE;', 'FALSE'))
                                  for inst in instances)
                total_force = sum(sum(1 for v in inst['dcc_config']['f_exps'].values() if v and v not in ('FALSE;', 'FALSE'))
                                   for inst in instances)
                total_at = sum(sum(1 for n, v in inst['dcc_config']['t_exps'].items()
                                   if v and v not in ('FALSE;', 'FALSE') and not inst['dcc_config']['t_disable'].get(n, False))
                               for inst in instances)
                self._log(f'[2/3] 解析完成:')
                self._log(f'  模块实例: {len(instances)} | 联锁: {total_ilock} | 允许: {total_perm} | 强制: {total_force} | 模拟跟踪: {total_at}')
                self._log(f'[3/3] 生成Excel: {os.path.basename(out)}')
                self.progress_var.set('正在生成Excel文件...')
                generate_excel(instances, out)
                self._log(f'✅ 完成! 文件已保存至: {out}')
                self.progress_var.set('完成!')
                self.status_var.set(f'完成 - {len(instances)}个模块, {total_ilock+total_perm+total_force+total_at}个条件')
                messagebox.showinfo('完成', f'联锁信息表已生成!\n\n{out}')
            except Exception as e:
                self._log(f'❌ 错误: {str(e)}')
                self.status_var.set('出错')
                messagebox.showerror('错误', str(e))
            finally:
                self._stop_progress()
                self.run_btn.configure(state='normal', text='▶ 开始提取')

        threading.Thread(target=worker, daemon=True).start()


if __name__ == '__main__':
    app = App()
    app.mainloop()
