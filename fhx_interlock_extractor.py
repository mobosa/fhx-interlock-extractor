#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
FHX Interlock Extractor
从DeltaV FHX文件中提取DCC功能块的联锁(Interlock)、允许(Permissive)、强制(Force)信息
并生成Excel联锁信息表

关键格式说明:
  MODULE_INSTANCE条目中使用 "DCC1$参数名" 格式（$分隔）存储实际配置
  MODULE_CLASS定义中使用 "DCC1/参数名" 格式（/分隔）存储默认值
  实际联锁配置在MODULE_INSTANCE的DCC1$I_EXP/DCC1$I_DESC/DCC1$I_STATE中
"""

import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def read_fhx_file(filepath):
    """读取FHX文件，处理UTF-16编码"""
    with open(filepath, 'rb') as f:
        raw = f.read()
    if raw[:2] == b'\xff\xfe':
        text = raw.decode('utf-16-le')
    elif raw[:2] == b'\xfe\xff':
        text = raw.decode('utf-16-be')
    else:
        text = raw.decode('utf-16', errors='ignore')
    return text.replace('\x00', '')


def extract_value(lines, start):
    """从当前位置提取ATTRIBUTE_INSTANCE的值（支持多行表达式）"""
    for i in range(start, min(start + 15, len(lines))):
        line = lines[i].strip()
        # 字符串值 CV="..."  (单行)
        cv = re.search(r'CV="([^"]*)"', line)
        if cv:
            return cv.group(1)
        # 整数值 CV=N
        cv_int = re.search(r'CV=(\d+)', line)
        if cv_int:
            return int(cv_int.group(1))
        # 表达式 EXPRESSION="..." - 可能跨多行
        # 找到 EXPRESSION=" 的位置
        exp_marker = 'EXPRESSION="'
        exp_pos = line.find(exp_marker)
        if exp_pos >= 0:
            # 提取EXPRESSION="之后的内容
            after = line[exp_pos + len(exp_marker):]
            # 检查当前行是否有闭合引号（排除字符串字面量中的引号）
            # 方法：去掉字符串字面量中的引号后检查
            # DeltaV表达式中字符串用单引号包围，所以双引号只在最外层
            if '"' in after:
                # 单行表达式 - 提取引号内的内容
                # 格式: EXPRESSION='...' }  或  EXPRESSION="..." }
                # 找到最后一个 " 的位置，它就是表达式的结束
                last_quote = after.rfind('"')
                if last_quote >= 0:
                    return after[:last_quote]
                return after
            else:
                # 多行表达式，继续读取直到找到表达式的闭合引号
                full_val = after
                for j in range(i + 1, min(i + 10, len(lines))):
                    next_line = lines[j].strip()
                    full_val += '\n' + next_line
                    # 检查是否到达表达式闭合引号
                    # FHX格式: 值" }  或  值"}
                    # 找到行内的 " 位置（表达式结束引号）
                    if '"' in next_line:
                        # 找到闭合引号
                        quote_pos = next_line.rfind('"')
                        # 取到引号之前的内容
                        before_quote = next_line[:quote_pos]
                        # 加到full_val并去掉尾部
                        full_val = full_val[:full_val.rfind('\n' + next_line)] + '\n' + before_quote
                        return full_val.strip()
                return full_val.strip()
        # 枚举 STRING_VALUE="..."
        sv = re.search(r'STRING_VALUE="([^"]*)"', line)
        if sv:
            return sv.group(1)
    return None


def parse_fhx(filepath):
    """解析FHX文件"""
    text = read_fhx_file(filepath)
    lines = text.split('\n')

    # 存储所有MODULE_INSTANCE的DCC配置
    instances = []  # [{tag, plant_area, description, module_class, dcc_config}]

    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # ========== 解析 MODULE_INSTANCE ==========
        mi_match = re.match(
            r'MODULE_INSTANCE\s+TAG="([^"]*)"\s+PLANT_AREA="([^"]*)"\s+MODULE_CLASS="([^"]*)"',
            line
        )
        if mi_match:
            tag = mi_match.group(1)
            plant_area = mi_match.group(2)
            mc_name = mi_match.group(3)

            # 找块体范围
            brace_count = 0
            block_start = i
            block_end = i
            found_open = False
            for j in range(i, min(i + 5000, len(lines))):
                for ch in lines[j]:
                    if ch == '{':
                        brace_count += 1
                        found_open = True
                    elif ch == '}':
                        brace_count -= 1
                if found_open and brace_count <= 0:
                    block_end = j
                    break

            block_lines = lines[block_start:block_end + 1]

            # 提取描述
            desc = ''
            for bl in block_lines:
                dm = re.search(r'DESCRIPTION="([^"]*)"', bl.strip())
                if dm:
                    desc = dm.group(1)
                    break

            # 检查是否有DCC1$或AT块开头的属性
            dcc_config = {
                'i_exps': {}, 'i_descs': {}, 'i_states': {},
                'i_disable': {}, 'i_higher_mng': {}, 'i_reset_reqd': {},
                'i_delay_on': {}, 'i_delay_off': {}, 'i_used': 0,
                'p_exps': {}, 'p_descs': {}, 'p_disable': {},
                'p_delay_on': {}, 'p_delay_off': {}, 'p_used': 0,
                'f_exps': {}, 'f_descs': {}, 'f_states': {},
                'f_disable': {}, 'f_delay_on': {}, 'f_used': 0,
                't_exps': {}, 't_descs': {}, 't_states': {},
                't_disable': {}, 't_reset_reqd': {}, 't_hold_man': {},
                't_delay_on': {}, 't_delay_off': {}, 't_val': {}, 't_used': 0,
                'block_type': '',
            }
            has_dcc_config = False

            for bi, bl in enumerate(block_lines):
                stripped = bl.strip()
                if 'ATTRIBUTE_INSTANCE' not in stripped:
                    continue

                # 只匹配 DCC1$ 或 AT1$/SM_AT1$/HM_AT1$ 等AT块
                attr_match = re.search(r'ATTRIBUTE_INSTANCE\s+NAME="(DCC1|AT\d+|SM_AT\d+|HM_AT\d+)\$(.+)"', stripped)
                if not attr_match:
                    continue

                block_name = attr_match.group(1)
                attr_name = attr_match.group(2)
                val = extract_value(block_lines, bi + 1)
                if val is None:
                    continue

                has_dcc_config = True

                # 判断块类型
                if attr_name.startswith(('I_', 'P_', 'F_')):
                    dcc_config['block_type'] = 'DCC'
                elif attr_name.startswith('T_'):
                    dcc_config['block_type'] = 'AT'

                # 联锁表达式
                m = re.match(r'I_EXP(\d+)$', attr_name)
                if m:
                    dcc_config['i_exps'][int(m.group(1))] = str(val)
                    continue

                # 联锁描述
                m = re.match(r'I_DESC_(\d+)$', attr_name)
                if m:
                    dcc_config['i_descs'][int(m.group(1))] = str(val)
                    continue

                # 联锁状态
                m = re.match(r'I_STATE(\d+)$', attr_name)
                if m:
                    dcc_config['i_states'][int(m.group(1))] = str(val)
                    continue

                # 联锁已用数量
                if attr_name == 'I_USED_CND':
                    dcc_config['i_used'] = int(val) if isinstance(val, int) else 0
                    continue

                # 联锁禁用
                m = re.match(r'I_DISABLE(\d+)$', attr_name)
                if m:
                    dcc_config['i_disable'][int(m.group(1))] = str(val).upper() == 'TRUE'
                    continue

                # 联锁高级管理
                m = re.match(r'I_HIGHER_MNG(\d+)$', attr_name)
                if m:
                    dcc_config['i_higher_mng'][int(m.group(1))] = str(val).upper() == 'TRUE'
                    continue

                # 联锁需要复位
                m = re.match(r'I_RESET_REQD(\d+)$', attr_name)
                if m:
                    dcc_config['i_reset_reqd'][int(m.group(1))] = str(val).upper() == 'TRUE'
                    continue

                # 联锁接通延时
                m = re.match(r'I_DELAY_ON(\d+)$', attr_name)
                if m:
                    dcc_config['i_delay_on'][int(m.group(1))] = val
                    continue

                # 联锁关断延时
                m = re.match(r'I_DELAY_OFF(\d+)$', attr_name)
                if m:
                    dcc_config['i_delay_off'][int(m.group(1))] = val
                    continue

                # 允许表达式
                m = re.match(r'P_EXP(\d+)$', attr_name)
                if m:
                    dcc_config['p_exps'][int(m.group(1))] = str(val)
                    continue

                # 允许描述
                m = re.match(r'P_DESC(\d+)$', attr_name)
                if m:
                    dcc_config['p_descs'][int(m.group(1))] = str(val)
                    continue

                # 允许已用数量
                if attr_name == 'P_USED_CND':
                    dcc_config['p_used'] = int(val) if isinstance(val, int) else 0
                    continue

                # 允许定时器
                m = re.match(r'P_TIMER(\d+)$', attr_name)
                if m:
                    dcc_config['p_timer'][int(m.group(1))] = val
                    continue

                # 允许禁用
                m = re.match(r'P_DISABLE(\d+)$', attr_name)
                if m:
                    dcc_config['p_disable'][int(m.group(1))] = str(val).upper() == 'TRUE'
                    continue

                # 允许接通延时
                m = re.match(r'P_DELAY_ON(\d+)$', attr_name)
                if m:
                    dcc_config['p_delay_on'][int(m.group(1))] = val
                    continue

                # 允许关断延时
                m = re.match(r'P_DELAY_OFF(\d+)$', attr_name)
                if m:
                    dcc_config['p_delay_off'][int(m.group(1))] = val
                    continue

                # 强制表达式
                m = re.match(r'F_EXP(\d+)$', attr_name)
                if m:
                    dcc_config['f_exps'][int(m.group(1))] = str(val)
                    continue

                # 强制描述
                m = re.match(r'F_DESC(\d+)$', attr_name)
                if m:
                    dcc_config['f_descs'][int(m.group(1))] = str(val)
                    continue

                # 强制状态
                m = re.match(r'F_STATE(\d+)$', attr_name)
                if m:
                    dcc_config['f_states'][int(m.group(1))] = str(val)
                    continue

                # 强制已用数量
                if attr_name == 'F_USED_CND':
                    dcc_config['f_used'] = int(val) if isinstance(val, int) else 0
                    continue

                # 强制定时器
                m = re.match(r'F_TIMER(\d+)$', attr_name)
                if m:
                    dcc_config['f_timer'][int(m.group(1))] = val
                    continue

                # 强制禁用
                m = re.match(r'F_DISABLE(\d+)$', attr_name)
                if m:
                    dcc_config['f_disable'][int(m.group(1))] = str(val).upper() == 'TRUE'
                    continue

                # 强制接通延时
                m = re.match(r'F_DELAY_ON(\d+)$', attr_name)
                if m:
                    dcc_config['f_delay_on'][int(m.group(1))] = val
                    continue

                # ========== AT 模拟跟踪参数 ==========
                m = re.match(r'T_EXP(\d+)$', attr_name)
                if m:
                    dcc_config['t_exps'][int(m.group(1))] = str(val)
                    continue
                m = re.match(r'T_DESC(\d+)$', attr_name)
                if m:
                    dcc_config['t_descs'][int(m.group(1))] = str(val)
                    continue
                m = re.match(r'T_STATE(\d+)$', attr_name)
                if m:
                    dcc_config['t_states'][int(m.group(1))] = str(val)
                    continue
                if attr_name == 'T_USED_CND':
                    dcc_config['t_used'] = int(val) if isinstance(val, int) else 0
                    continue
                m = re.match(r'T_DISABLE(\d+)$', attr_name)
                if m:
                    dcc_config['t_disable'][int(m.group(1))] = str(val).upper() == 'TRUE'
                    continue
                m = re.match(r'T_RESET_REQD(\d+)$', attr_name)
                if m:
                    dcc_config['t_reset_reqd'][int(m.group(1))] = str(val).upper() == 'TRUE'
                    continue
                m = re.match(r'T_HOLD_MAN(\d+)$', attr_name)
                if m:
                    dcc_config['t_hold_man'][int(m.group(1))] = str(val).upper() == 'TRUE'
                    continue
                m = re.match(r'T_DELAY_ON(\d+)$', attr_name)
                if m:
                    dcc_config['t_delay_on'][int(m.group(1))] = val
                    continue
                m = re.match(r'T_DELAY_OFF(\d+)$', attr_name)
                if m:
                    dcc_config['t_delay_off'][int(m.group(1))] = val
                    continue
                m = re.match(r'T_VAL(\d+)$', attr_name)
                if m:
                    dcc_config['t_val'][int(m.group(1))] = val
                    continue

            if has_dcc_config:
                # 检查是否有实际配置的条件（排除全为FALSE的）
                dc = dcc_config
                has_real = False
                for exp_dict in [dc['i_exps'], dc['p_exps'], dc['f_exps'], dc['t_exps']]:
                    for v in exp_dict.values():
                        if v and v not in ('FALSE;', 'FALSE'):
                            has_real = True
                            break
                    if has_real:
                        break
                if has_real:
                    instances.append({
                        'tag': tag,
                        'plant_area': plant_area,
                        'description': desc,
                        'module_class': mc_name,
                        'dcc_config': dcc_config,
                    })

            i = block_end + 1
            continue

        i += 1

    return instances


def generate_excel(instances, output_path):
    """生成Excel联锁信息表"""
    wb = Workbook()

    # 样式
    hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    hdr_fill_blue = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hdr_fill_red = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    hdr_fill_green = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
    hdr_fill_gold = PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid')
    nfont = Font(name='微软雅黑', size=10)
    c_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    l_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fill_ilock = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    fill_perm = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    fill_force = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')

    def set_header(ws, row, headers, fill):
        for col, (text, width) in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=text)
            c.font = hdr_font
            c.fill = fill
            c.alignment = c_align
            c.border = border
            ws.column_dimensions[get_column_letter(col)].width = width

    def sc(ws, row, col, value, font=nfont, align=c_align, fill=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font
        c.alignment = align
        c.border = border
        if fill:
            c.fill = fill
        return c

    # ========== Sheet 1: 联锁信息总表 ==========
    ws1 = wb.active
    ws1.title = "联锁信息总表"
    ws1.merge_cells('A1:O1')
    t = ws1['A1']
    t.value = 'DeltaV DCC 联锁信息总表'
    t.font = Font(name='微软雅黑', bold=True, size=16, color='1F4E79')
    t.alignment = Alignment(horizontal='center', vertical='center')

    h1 = [
        ('序号', 6), ('TAG', 14), ('Plant Area', 20), ('Description', 22), ('Module Class', 18),
        ('类型', 8), ('Condition No', 8), ('Description', 22), ('Condition Expression', 48),
        ('Interlock State', 12), ('Force SP Value', 13), ('Disable', 7), ('Higher Managed', 12),
        ('Reset Required', 11), ('Delay On (s)', 10), ('Delay Off (s)', 10),
        ('Used(I)', 9), ('Used(P)', 9), ('Used(F)', 9),
    ]
    set_header(ws1, 2, h1, hdr_fill_blue)

    row = 3
    seq = 1
    for inst in sorted(instances, key=lambda x: x['tag']):
        dc = inst['dcc_config']

        # 收集所有条件
        all_conds = []

        # 联锁
        for num in sorted(dc['i_exps'].keys()):
            exp = dc['i_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE') and not dc['i_disable'].get(num, False):
                desc = dc['i_descs'].get(num, '')
                state = dc['i_states'].get(num, '')
                disable = dc['i_disable'].get(num, False)
                higher = dc['i_higher_mng'].get(num, False)
                reset = dc['i_reset_reqd'].get(num, False)
                delay_on = dc['i_delay_on'].get(num, 0)
                delay_off = dc['i_delay_off'].get(num, 0)
                all_conds.append(('联锁', num, desc, exp, state, disable, higher, reset, delay_on, delay_off))

        # 允许
        for num in sorted(dc['p_exps'].keys()):
            exp = dc['p_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE'):
                desc = dc['p_descs'].get(num, '')
                all_conds.append(('允许', num, desc, exp, '', False, False, False, 0, 0))

        # 强制
        for num in sorted(dc['f_exps'].keys()):
            exp = dc['f_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE'):
                desc = dc['f_descs'].get(num, '')
                state = dc['f_states'].get(num, '')
                all_conds.append(('强制', num, desc, exp, state, False, False, False, 0, 0))

        # 模拟跟踪
        for num in sorted(dc['t_exps'].keys()):
            exp = dc['t_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE') and not dc['t_disable'].get(num, False):
                desc = dc['t_descs'].get(num, '')
                disable = dc['t_disable'].get(num, False)
                hold_man = dc['t_hold_man'].get(num, False)
                reset = dc['t_reset_reqd'].get(num, False)
                delay_on = dc['t_delay_on'].get(num, 0)
                delay_off = dc['t_delay_off'].get(num, 0)
                all_conds.append(('模拟跟踪', num, desc, exp, '', disable, False, reset, delay_on, delay_off))

        if not all_conds:
            sc(ws1, row, 1, seq)
            sc(ws1, row, 2, inst['tag'])
            sc(ws1, row, 3, inst['plant_area'], align=l_align)
            sc(ws1, row, 4, inst['description'], align=l_align)
            sc(ws1, row, 5, inst['module_class'], align=l_align)
            sc(ws1, row, 6, '-')
            sc(ws1, row, 7, '-')
            sc(ws1, row, 8, '(无配置)', align=l_align)
            sc(ws1, row, 9, '-', align=l_align)
            sc(ws1, row, 10, '-')
            sc(ws1, row, 11, '-')
            sc(ws1, row, 12, '-')
            sc(ws1, row, 13, '-')
            sc(ws1, row, 14, 0)
            sc(ws1, row, 15, 0)
            sc(ws1, row, 16, dc['i_used'])
            sc(ws1, row, 17, dc['p_used'])
            sc(ws1, row, 18, dc['f_used'])
            row += 1
            seq += 1
        else:
            first = True
            cond_seq = 0
            for ctype, cnum, cdesc, cexp, cstate, cdisable, chigher, creset, cdelay_on, cdelay_off in all_conds:
                cond_seq += 1
                sc(ws1, row, 1, seq if first else '')
                sc(ws1, row, 2, inst['tag'] if first else '')
                sc(ws1, row, 3, inst['plant_area'] if first else '', align=l_align)
                sc(ws1, row, 4, inst['description'] if first else '', align=l_align)
                sc(ws1, row, 5, inst['module_class'] if first else '', align=l_align)
                fill_map = {'联锁': fill_ilock, '允许': fill_perm, '强制': fill_force}
                fill_at = PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid')
                tc = sc(ws1, row, 6, ctype, fill=fill_map.get(ctype, fill_at))
                sc(ws1, row, 7, cond_seq)
                sc(ws1, row, 8, cdesc, align=l_align)
                sc(ws1, row, 9, cexp, align=l_align)
                # Interlock State (联锁用) / Force SP Value (强制用)
                sc(ws1, row, 10, cstate if ctype == '联锁' else '')
                sc(ws1, row, 11, cstate if ctype == '强制' else '')
                sc(ws1, row, 12, '是' if cdisable else '否')
                sc(ws1, row, 13, '是' if chigher else '否')
                sc(ws1, row, 14, '是' if creset else '否')
                sc(ws1, row, 15, cdelay_on)
                sc(ws1, row, 16, cdelay_off)
                sc(ws1, row, 17, dc['i_used'] if first else '')
                sc(ws1, row, 18, dc['p_used'] if first else '')
                sc(ws1, row, 19, dc['f_used'] if first else '')
                row += 1
                first = False
            seq += 1

    ws1.freeze_panes = 'A3'
    ws1.auto_filter.ref = f'A2:S{row - 1}'

    # ========== Sheet 2: 联锁条件明细 ==========
    ws2 = wb.create_sheet("联锁条件明细")
    ws2.merge_cells('A1:H1')
    ws2['A1'].value = '联锁条件明细表'
    ws2['A1'].font = Font(name='微软雅黑', bold=True, size=14, color='C00000')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')

    h2 = [('序号', 6), ('TAG', 14), ('Plant Area', 18), ('Condition No', 8), ('Description', 22),
          ('Condition Expression', 48), ('Interlock State', 12), ('Disable', 7),
          ('Higher Managed', 12), ('Reset Required', 11), ('Delay On (s)', 10),
          ('Delay Off (s)', 10), ('Used/Max', 10)]
    set_header(ws2, 2, h2, hdr_fill_red)

    row = 3
    seq = 1
    for inst in sorted(instances, key=lambda x: x['tag']):
        dc = inst['dcc_config']
        for num in sorted(dc['i_exps'].keys()):
            exp = dc['i_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE'):
                sc(ws2, row, 1, seq)
                sc(ws2, row, 2, inst['tag'])
                sc(ws2, row, 3, inst['plant_area'], align=l_align)
                sc(ws2, row, 4, seq)
                sc(ws2, row, 5, dc['i_descs'].get(num, ''), align=l_align)
                sc(ws2, row, 6, exp, align=l_align)
                sc(ws2, row, 7, dc['i_states'].get(num, ''))
                sc(ws2, row, 8, '是' if dc['i_disable'].get(num, False) else '否')
                sc(ws2, row, 9, '是' if dc['i_higher_mng'].get(num, False) else '否')
                sc(ws2, row, 10, '是' if dc['i_reset_reqd'].get(num, False) else '否')
                sc(ws2, row, 11, dc['i_delay_on'].get(num, 0))
                sc(ws2, row, 12, dc['i_delay_off'].get(num, 0))
                sc(ws2, row, 13, f"{dc['i_used']}/16")
                row += 1
                seq += 1

    ws2.freeze_panes = 'A3'
    if row > 3:
        ws2.auto_filter.ref = f'A2:M{row - 1}'

    # ========== Sheet 3: 允许条件明细 ==========
    ws3 = wb.create_sheet("允许条件明细")
    ws3.merge_cells('A1:G1')
    ws3['A1'].value = '允许条件明细表'
    ws3['A1'].font = Font(name='微软雅黑', bold=True, size=14, color='548235')
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')

    h3 = [('序号', 6), ('TAG', 14), ('Plant Area', 18), ('Condition No', 8), ('Description', 22),
          ('Condition Expression', 48), ('Disable', 7), ('Delay On (s)', 10),
          ('Delay Off (s)', 10), ('Used/Max', 10)]
    set_header(ws3, 2, h3, hdr_fill_green)

    row = 3
    seq = 1
    for inst in sorted(instances, key=lambda x: x['tag']):
        dc = inst['dcc_config']
        for num in sorted(dc['p_exps'].keys()):
            exp = dc['p_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE'):
                sc(ws3, row, 1, seq)
                sc(ws3, row, 2, inst['tag'])
                sc(ws3, row, 3, inst['plant_area'], align=l_align)
                sc(ws3, row, 4, seq)
                sc(ws3, row, 5, dc['p_descs'].get(num, ''), align=l_align)
                sc(ws3, row, 6, exp, align=l_align)
                sc(ws3, row, 7, '是' if dc['p_disable'].get(num, False) else '否')
                sc(ws3, row, 8, dc['p_delay_on'].get(num, 0))
                sc(ws3, row, 9, dc['p_delay_off'].get(num, 0))
                sc(ws3, row, 10, f"{dc['p_used']}/8")
                row += 1
                seq += 1

    ws3.freeze_panes = 'A3'
    if row > 3:
        ws3.auto_filter.ref = f'A2:J{row - 1}'

    # ========== Sheet 4: 强制条件明细 ==========
    ws4 = wb.create_sheet("强制条件明细")
    ws4.merge_cells('A1:J1')
    ws4['A1'].value = '强制设定点条件明细表'
    ws4['A1'].font = Font(name='微软雅黑', bold=True, size=14, color='BF8F00')
    ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')

    h4 = [('序号', 6), ('TAG', 14), ('Plant Area', 18), ('Condition No', 8), ('Description', 22),
          ('Condition Expression', 48), ('Force SP Value', 12), ('Disable', 7),
          ('Delay On (s)', 10), ('Used/Max', 10)]
    set_header(ws4, 2, h4, hdr_fill_gold)

    row = 3
    seq = 1
    for inst in sorted(instances, key=lambda x: x['tag']):
        dc = inst['dcc_config']
        for num in sorted(dc['f_exps'].keys()):
            exp = dc['f_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE'):
                sc(ws4, row, 1, seq)
                sc(ws4, row, 2, inst['tag'])
                sc(ws4, row, 3, inst['plant_area'], align=l_align)
                sc(ws4, row, 4, seq)
                sc(ws4, row, 5, dc['f_descs'].get(num, ''), align=l_align)
                sc(ws4, row, 6, exp, align=l_align)
                sc(ws4, row, 7, dc['f_states'].get(num, ''))
                sc(ws4, row, 8, '是' if dc['f_disable'].get(num, False) else '否')
                sc(ws4, row, 9, dc['f_delay_on'].get(num, 0))
                sc(ws4, row, 10, f"{dc['f_used']}/8")
                row += 1
                seq += 1

    ws4.freeze_panes = 'A3'
    if row > 3:
        ws4.auto_filter.ref = f'A2:J{row - 1}'

    # ========== Sheet 5: 模拟跟踪条件明细 ==========
    ws_at = wb.create_sheet("模拟跟踪条件明细")
    ws_at.merge_cells('A1:L1')
    ws_at['A1'].value = 'Analog Tracking (AT) 条件明细表'
    ws_at['A1'].font = Font(name='微软雅黑', bold=True, size=14, color='7030A0')
    ws_at['A1'].alignment = Alignment(horizontal='center', vertical='center')

    hdr_fill_purple = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
    h_at = [('序号', 6), ('TAG', 14), ('Plant Area', 18), ('Condition No', 8), ('Description', 22),
            ('Condition Expression', 48), ('Track Value', 10), ('Disable', 7),
            ('Hold in Manual', 13), ('Reset Required', 11), ('Delay On (s)', 10),
            ('Delay Off (s)', 10), ('Used/Max', 10)]
    set_header(ws_at, 2, h_at, hdr_fill_purple)

    row = 3
    seq = 1
    for inst in sorted(instances, key=lambda x: x['tag']):
        dc = inst['dcc_config']
        for num in sorted(dc['t_exps'].keys()):
            exp = dc['t_exps'][num]
            if exp and exp not in ('FALSE;', 'FALSE') and not dc['t_disable'].get(num, False):
                sc(ws_at, row, 1, seq)
                sc(ws_at, row, 2, inst['tag'])
                sc(ws_at, row, 3, inst['plant_area'], align=l_align)
                sc(ws_at, row, 4, seq)
                sc(ws_at, row, 5, dc['t_descs'].get(num, ''), align=l_align)
                sc(ws_at, row, 6, exp, align=l_align)
                sc(ws_at, row, 7, dc['t_val'].get(num, 0))
                sc(ws_at, row, 8, '是' if dc['t_disable'].get(num, False) else '否')
                sc(ws_at, row, 9, '是' if dc['t_hold_man'].get(num, False) else '否')
                sc(ws_at, row, 10, '是' if dc['t_reset_reqd'].get(num, False) else '否')
                sc(ws_at, row, 11, dc['t_delay_on'].get(num, 0))
                sc(ws_at, row, 12, dc['t_delay_off'].get(num, 0))
                sc(ws_at, row, 13, f"{dc['t_used']}/16")
                row += 1
                seq += 1

    ws_at.freeze_panes = 'A3'
    if row > 3:
        ws_at.auto_filter.ref = f'A2:M{row - 1}'

    # ========== Sheet 6: 所有模块实例汇总 ==========
    ws5 = wb.create_sheet("模块实例汇总")
    ws5.merge_cells('A1:G1')
    ws5['A1'].value = '含联锁配置的模块实例汇总'
    ws5['A1'].font = Font(name='微软雅黑', bold=True, size=14, color='1F4E79')
    ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')

    h5 = [('序号', 6), ('TAG', 14), ('区域', 20), ('描述', 22), ('模块类', 18), ('联锁数', 10), ('允许数', 10)]
    set_header(ws5, 2, h5, hdr_fill_blue)

    row = 3
    seq = 1
    for inst in sorted(instances, key=lambda x: x['tag']):
        dc = inst['dcc_config']
        ilock_count = sum(1 for v in dc['i_exps'].values() if v and v not in ('FALSE;', 'FALSE'))
        perm_count = sum(1 for v in dc['p_exps'].values() if v and v not in ('FALSE;', 'FALSE'))
        force_count = sum(1 for v in dc['f_exps'].values() if v and v not in ('FALSE;', 'FALSE'))

        sc(ws5, row, 1, seq)
        sc(ws5, row, 2, inst['tag'])
        sc(ws5, row, 3, inst['plant_area'], align=l_align)
        sc(ws5, row, 4, inst['description'], align=l_align)
        sc(ws5, row, 5, inst['module_class'], align=l_align)
        sc(ws5, row, 6, ilock_count)
        sc(ws5, row, 7, perm_count)
        row += 1
        seq += 1

    ws5.freeze_panes = 'A3'
    ws5.auto_filter.ref = f'A2:G{row - 1}'

    wb.save(output_path)
    return output_path


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    fhx_path = os.path.join(base_dir, 'fault.fhx')
    output_path = os.path.join(base_dir, '联锁信息表.xlsx')

    if not os.path.exists(fhx_path):
        print(f"错误: 找不到FHX文件 {fhx_path}")
        return

    print(f"[1/3] 读取FHX文件: {fhx_path}")
    print(f"      文件大小: {os.path.getsize(fhx_path) / 1024:.1f} KB")

    print(f"[2/3] 解析DCC联锁配置...")
    instances = parse_fhx(fhx_path)

    # 统计
    total_ilock = sum(
        sum(1 for num, v in inst['dcc_config']['i_exps'].items()
            if v and v not in ('FALSE;', 'FALSE') and not inst['dcc_config']['i_disable'].get(num, False))
        for inst in instances
    )
    total_perm = sum(
        sum(1 for v in inst['dcc_config']['p_exps'].values() if v and v not in ('FALSE;', 'FALSE'))
        for inst in instances
    )
    total_force = sum(
        sum(1 for v in inst['dcc_config']['f_exps'].values() if v and v not in ('FALSE;', 'FALSE'))
        for inst in instances
    )
    total_at = sum(
        sum(1 for num, v in inst['dcc_config']['t_exps'].items()
            if v and v not in ('FALSE;', 'FALSE') and not inst['dcc_config']['t_disable'].get(num, False))
        for inst in instances
    )

    print(f"      含DCC/AT配置的模块实例: {len(instances)} 个")
    print(f"      已配置联锁条件: {total_ilock} 个")
    print(f"      已配置允许条件: {total_perm} 个")
    print(f"      已配置强制条件: {total_force} 个")
    print(f"      已配置模拟跟踪条件: {total_at} 个")

    print(f"[3/3] 生成Excel文件: {output_path}")
    generate_excel(instances, output_path)
    print(f"\n完成! Excel文件已保存至: {output_path}")


if __name__ == '__main__':
    main()
